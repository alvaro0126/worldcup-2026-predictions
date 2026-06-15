"""Write the Excel workbook with all sheets, lightly formatted."""
from __future__ import annotations
import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from . import config
from .data import Team, Match, Result
from .bayes import BayesModel
from .simulate import SimAgg
from .match_model import Params
from .recommend import Rec, group_recs, knockout_recs
from .metrics import Validation

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _score(h: int, a: int) -> str:
    return f"{int(h)}-{int(a)}"


def _rec_row(r: Rec) -> dict:
    return {
        "Match": r.match_id,
        "Round": config.ROUND_NAMES.get(r.rnd, r.rnd),
        "Group": r.group or "",
        "Date": r.date,
        "Home": r.home,
        "Away": r.away,
        "REC SCORE": _score(r.rec_home, r.rec_away),
        "Exp pts": round(r.ev),
        "Status": r.status,
        "Win%": round(r.p_home * 100, 1),
        "Draw%": round(r.p_draw * 100, 1),
        "Loss%": round(r.p_away * 100, 1),
        "xG home": round(r.exp_home, 2),
        "xG away": round(r.exp_away, 2),
        "Pick": {"H": "Home", "D": "Draw", "A": "Away"}[r.best_outcome],
        "Most likely": _score(r.alt_home, r.alt_away),
        "Actual": "" if r.actual_home is None else _score(r.actual_home, r.actual_away),
        "Matchup prob": "" if r.matchup_prob is None else round(r.matchup_prob * 100, 1),
        "Note": r.note,
    }


def _team_table(bm: BayesModel, teams: dict[str, Team], agg: SimAgg) -> pd.DataFrame:
    n = agg.n
    rows = []
    for name in sorted(teams, key=lambda t: bm.effective_elo(t), reverse=True):
        t = teams[name]
        prior, post = bm.prior_elo(name), bm.effective_elo(name)
        reached = agg.reached[name]
        rows.append({
            "Team": name, "Group": t.group, "Conf": t.confederation,
            "Elo (prior)": round(prior), "Elo (updated)": round(post),
            "Shift": round(post - prior),
            "Elo source": t.elo_source,
            "FIFA rank": "" if t.fifa_rank is None else t.fifa_rank,
            "Expert rating": "" if t.expert_rating is None else t.expert_rating,
            "Games": int(bm.ngames[bm.idx[name]]),
            "Win title %": round(agg.champion[name] / n * 100, 1),
            "Reach final %": round((agg.champion[name] + agg.runner_up[name]) / n * 100, 1),
            "Advance %": round(agg.advanced[name] / n * 100, 1),
        })
    return pd.DataFrame(rows)


def _title_table(bm: BayesModel, teams: dict[str, Team], agg: SimAgg) -> pd.DataFrame:
    n = agg.n
    rows = []
    for name in teams:
        rc = agg.reached[name]
        reach_r16 = (rc["R16"] + rc["QF"] + rc["SF"] + rc["F"] + rc["W"]) / n
        reach_qf = (rc["QF"] + rc["SF"] + rc["F"] + rc["W"]) / n
        reach_sf = (rc["SF"] + rc["F"] + rc["W"]) / n
        reach_f = (rc["F"] + rc["W"]) / n
        rows.append({
            "Team": name, "Group": teams[name].group,
            "Advance %": round(agg.advanced[name] / n * 100, 1),
            "Reach QF %": round(reach_qf * 100, 1),
            "Reach SF %": round(reach_sf * 100, 1),
            "Reach final %": round(reach_f * 100, 1),
            "Win title %": round(agg.champion[name] / n * 100, 1),
        })
    df = pd.DataFrame(rows).sort_values("Win title %", ascending=False)
    return df.reset_index(drop=True)


def _standings_table(teams: dict[str, Team], agg: SimAgg) -> pd.DataFrame:
    n = agg.n
    rows = []
    for name in teams:
        gp = agg.group_pos[name]
        rows.append({
            "Group": teams[name].group, "Team": name,
            "P(1st) %": round(gp[0] / n * 100, 1),
            "P(2nd) %": round(gp[1] / n * 100, 1),
            "P(3rd) %": round(gp[2] / n * 100, 1),
            "P(4th) %": round(gp[3] / n * 100, 1),
            "Advance %": round(agg.advanced[name] / n * 100, 1),
        })
    df = pd.DataFrame(rows)
    df = df.sort_values(["Group", "Advance %"], ascending=[True, False])
    return df.reset_index(drop=True)


def _readme_lines(params: Params, val: Validation, agg: SimAgg) -> list[list[str]]:
    top = agg.champion.most_common(3)
    return [
        ["World Cup 2026 — Score Recommendations"],
        [f"Generated from a Dixon-Coles bivariate-Poisson model with a daily Bayesian update.  {agg.n:,} tournament simulations."],
        [""],
        ["HOW TO READ"],
        ["• 'Recommendations (104)' is the headline sheet: one recommended exact score per match."],
        ["• Group matches use real fixtures; played games show the actual score. Knockout matches show the most-likely matchup for that bracket slot (flagged 'projected')."],
        ["• 'REC SCORE' maximises expected Bodytech points (140 exact / 100 winner+diff / 70 winner or draw). 'Exp pts' is its average points; 'Most likely' is the plain most-probable scoreline."],
        [""],
        ["MODEL (sound, reproducible)"],
        ["• Team strength = World Football Elo (eloratings.net), blended for lower teams with FIFA ranking + a composite expert rating."],
        [f"• Match engine = Dixon-Coles Poisson, calibrated to benchmark match probabilities (alpha={params.alpha:.3f}, b={params.b:.3f}, rho={params.rho:.3f})."],
        ["• Bayesian update: each FINAL result nudges a team's attack/defence, shrunk toward its Elo prior (one game can't overhaul a team)."],
        ["• Live/in-progress games are held out of the update until full-time."],
        [""],
        ["VALIDATION (pre-tournament prior vs games already played, honest hold-out)"],
        [f"• Matches scored: {val.n}   Outcome hit-rate: {val.outcome_hit_rate:.0%}   Mean RPS: {val.mean_rps:.3f}  (good tournament forecasts ~0.17-0.18)"],
        [""],
        ["CURRENT TOP-3 TITLE ODDS"],
        [f"• " + ",  ".join(f"{t} {c/agg.n:.0%}" for t, c in top)],
        [""],
        ["CAVEATS"],
        ["• A World Cup is designed to be unpredictable — even the favourite wins only ~1 in 6. These are probabilistic best guesses, not locks."],
        ["• Knockout matchups are projections; re-run after the group stage finishes to lock real matchups."],
        ["• Knockouts are treated as neutral-venue (host advantage only modelled in the group stage)."],
        ["• Squad-value data (Transfermarkt) is proxied via Elo, which lifts star-squad teams (France/Brazil) slightly."],
        ["", ],
        ["TO UPDATE DAILY"],
        ["• Edit data/results.csv (set status=final as games finish), then run:  python run.py"],
    ]


def _format_sheet(ws, df_like_header_row: int = 1, freeze: str = "A2") -> None:
    for cell in ws[df_like_header_row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = freeze
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 8), 42)


def build(path, bm: BayesModel, teams: dict[str, Team], fixtures: list[Match],
          results: list[Result], agg: SimAgg, params: Params,
          val: Validation) -> None:
    grecs = group_recs(bm, fixtures, results)
    krecs = knockout_recs(bm, agg)

    df_all = pd.DataFrame([_rec_row(r) for r in grecs + krecs])
    df_group = pd.DataFrame([_rec_row(r) for r in grecs])
    df_ko = pd.DataFrame([_rec_row(r) for r in krecs])
    df_teams = _team_table(bm, teams, agg)
    df_title = _title_table(bm, teams, agg)
    df_stand = _standings_table(teams, agg)
    df_val = pd.DataFrame(val.rows)

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        # read-me first
        pd.DataFrame(_readme_lines(params, val, agg)).to_excel(
            xw, sheet_name="Read me", index=False, header=False)
        df_all.to_excel(xw, sheet_name="Recommendations (104)", index=False)
        df_group.to_excel(xw, sheet_name="Group fixtures", index=False)
        df_ko.to_excel(xw, sheet_name="Projected knockouts", index=False)
        df_stand.to_excel(xw, sheet_name="Group finish odds", index=False)
        df_title.to_excel(xw, sheet_name="Title odds", index=False)
        df_teams.to_excel(xw, sheet_name="Team ratings", index=False)
        df_val.to_excel(xw, sheet_name="Validation", index=False)

        wb = xw.book
        for sheet in ["Recommendations (104)", "Group fixtures", "Projected knockouts",
                      "Group finish odds", "Title odds", "Team ratings", "Validation"]:
            _format_sheet(wb[sheet])
        # light touch on the read-me
        rm = wb["Read me"]
        rm["A1"].font = Font(bold=True, size=14, color="1F3864")
        rm.column_dimensions["A"].width = 120
